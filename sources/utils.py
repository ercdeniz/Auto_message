from datetime import date
import os
from datetime import datetime

import openpyxl
import pywhatkit as kit

import constants as const


# change this function according to your own Excel file
def get_data(file_path, sheet_name):
    excel_file = openpyxl.load_workbook(file_path)
    sheet = excel_file[sheet_name]
    total_rows = sheet.max_row
    data = []
    const.current_month = sheet.cell(2, 4).value
    const.next_month = sheet.cell(2, 5).value
    for row in range(2, total_rows + 1):
        name = sheet.cell(row, 1).value
        phone_number = sheet.cell(row, 2).value
        juz = sheet.cell(row, 3).value
        map_data = {"name": name, "phone_number": phone_number, "juz": juz}
        data.append(map_data)
    return data


def send_message(persons, index, day_left, message):
    message_text = ""
    if not persons[index]["juz"] == "-":
        if message == "reminder":
            message_text = const.message_reminder.format(
                month=const.current_month,
                name=persons[index]["name"],
                juz=persons[index]["juz"],
                day_left=day_left
            )
        elif message == "control":
            message_text = const.message_control.format(
                month=const.current_month,
                name=persons[index]["name"],
                juz=persons[index]["juz"]
            )
    if message == "request":
        message_text = const.message_request.format(
            month=const.next_month,
            name=persons[index]["name"]
        )
    if message_text:
        kit.sendwhatmsg_instantly(persons[index]["phone_number"], message_text, tab_close=True)


def day_calculator(year, month, day):
    future_date = date(year, month, day)
    current_date = date.today()
    day_left = (future_date - current_date).days
    return day_left


def logger(people, exception, file_name):
    error_message = (
        f"Message sending error:\n"
        f"   person:\n"
        f"       name: {people[const.index]['name']}\n"
        f"       number: {people[const.index]['phone_number']}\n"
        f"       juz: {people[const.index]['juz']}\n"
        f"   error: {exception}\n"
        f"--------------------\n"
    )
    with open(file_name, "a", encoding="utf-8") as log_file:
        log_file.write(error_message)


def move_file_with_timestamp(logs_directory):
    source_file = "PyWhatKit_DB.txt"
    try:
        if os.path.exists(source_file):
            timestamp = datetime.now().strftime("%d-%m_%H-%M-%S")
            destination_file = os.path.join(logs_directory, f"PyWhatKit_DB_{timestamp}.txt")
            os.rename(source_file, destination_file)
        else:
            raise FileNotFoundError(f"{source_file} not found.")
    except Exception as e:
        print(f"Error: {e}")
